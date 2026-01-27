import time
import os
import re
import pandas as pd
import openpyxl
from datetime import datetime
from playwright.sync_api import sync_playwright
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置
USERNAME = "伟填"
PASSWORD = "Test0528."
LOGIN_URL = "https://szguokuai.zlj.xyzulin.top/web/index.php?c=site&a=entry&m=ewei_shopv2&do=web&r=goods"
ID_RECORD_FILE = "processed_ids.txt"
MAX_PAGES = 0  # 最大抓取页数，设置 0 为不限制（抓取所有页）
HEADLESS = False
OUTPUT_FILE = f"scrape_goods_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
FORCE_UPDATE = True # 是否强制更新（忽略已处理记录，重新抓取所有数据）

def load_processed_ids():
    if os.path.exists(ID_RECORD_FILE):
        with open(ID_RECORD_FILE, "r", encoding="utf-8") as f:
            return set(line.strip() for line in f if line.strip())
    return set()

def save_new_ids(new_ids):
    with open(ID_RECORD_FILE, "a", encoding="utf-8") as f:
        for gid in new_ids:
            f.write(f"{gid}\n")

def update_master_headers(master_headers, current_headers):
    """
    合并新表头到主表头列表，保持相对顺序。
    """
    last_index = -1
    for header in current_headers:
        if header in master_headers:
            last_index = master_headers.index(header)
        else:
            # 插入到上一个已知列的后面
            insert_pos = last_index + 1
            master_headers.insert(insert_pos, header)
            last_index = insert_pos

def run_scraping():
    processed_ids = load_processed_ids()
    if FORCE_UPDATE:
        print("注意：强制更新模式已开启，将忽略历史记录重新抓取所有数据。")
        processed_ids = set() # 清空内存中的记录，视为全部新增
    else:
        print(f"已加载 {len(processed_ids)} 个历史商品ID。")
    
    all_sku_rows = []
    master_sku_headers = [] # 用于记录所有SKU列的正确顺序
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        
        # --- 登录阶段 ---
        print(f"正在访问登录页面: {LOGIN_URL}")
        page.goto(LOGIN_URL)
        try:
            page.wait_for_load_state('networkidle')
            if "login" in page.url or page.query_selector("input[type='password']"):
                print(f"检测到需要登录 (当前URL: {page.url})")
                
                if USERNAME and PASSWORD:
                    print(f"尝试自动登录...")
                    if page.query_selector("input[name='username']"):
                        page.fill("input[name='username']", USERNAME)
                    if page.query_selector("input[name='password']"):
                        page.fill("input[name='password']", PASSWORD)
                    submit_btn = page.query_selector("input[type='submit']") or page.query_selector("button[type='submit']") or page.query_selector(".btn-submit")
                    if submit_btn:
                        submit_btn.click()
                
                print("等待登录跳转...")
                try:
                    page.wait_for_selector("input[type='password']", state="hidden", timeout=5000)
                except:
                    pass
                
                page.wait_for_timeout(3000)
                print(f"当前页面 URL: {page.url}")

                if "r=goods" not in page.url or "login" in page.url:
                    print("未自动跳转到商品列表页，尝试强制访问...")
                    page.goto(LOGIN_URL)
                    page.wait_for_load_state('networkidle')
                
                print("登录流程结束。")
            else:
                print("已处于登录状态。")
                
            if page.query_selector("table"):
                print("成功检测到商品列表表格！")
            else:
                print("警告：当前页面未找到表格，可能需要手动干预。")
                
        except Exception as e:
            print(f"登录检查异常: {e}")

        # --- 第一阶段：扫描列表页收集新ID ---
        print("\n=== 第一阶段：扫描列表页收集新ID ===")
        ids_to_process = []
        page_num = 1
        
        # 如果开启强制更新，则不使用 processed_ids 进行过滤
        filter_ids = set() if FORCE_UPDATE else processed_ids
        
        while True:
            if MAX_PAGES > 0 and page_num > MAX_PAGES:
                print(f"已达到最大页数限制 ({MAX_PAGES})，停止扫描。")
                break

            print(f"正在扫描列表第 {page_num} 页...")
            try:
                page.wait_for_selector("table > tbody > tr", timeout=10000)
            except:
                print("未找到表格行，可能已无数据或加载超时。")
                break

            rows = page.query_selector_all("body > div.wb-container > div.page-content > div.row > div > table > tbody > tr")
            current_page_ids = []
            
            for row in rows:
                try:
                    id_cell = row.query_selector("td:nth-child(2)")
                    if id_cell:
                        goods_id = id_cell.inner_text().strip()
                        if goods_id.isdigit():
                            current_page_ids.append(goods_id)
                except:
                    pass
            
            # 使用 filter_ids (根据 FORCE_UPDATE 决定是否为空)
            new_ids_on_page = [gid for gid in current_page_ids if gid not in filter_ids and gid not in ids_to_process]
            
            print(f"  - 第 {page_num} 页共 {len(current_page_ids)} 个ID，其中新增: {len(new_ids_on_page)} 个")
            
            if new_ids_on_page:
                ids_to_process.extend(new_ids_on_page)
                # print(f"  - 已加入待抓取队列: {new_ids_on_page}") # 减少日志输出
            else:
                if current_page_ids:
                    print("  - 当前页未发现新ID，继续扫描...")
                    # 移除 break 以支持全量扫描
            
            user_next_selector = "ul.pagination > li > a[aria-label='Next']"
            next_btn = page.query_selector(user_next_selector)
            
            if not next_btn:
                 # 尝试其他可能的选择器，比如直接找 "下一页" 文本
                 next_btn = page.query_selector("ul.pagination li a:has-text('下一页')") or \
                            page.query_selector("ul.pagination li a:has-text('»')")
            
            # 有些时候下一页按钮虽然存在，但是是被禁用的 (disabled)，或者没有 href
            if next_btn:
                # 检查是否是 disabled
                # Playwright ElementHandle 没有 xpath 方法，使用 evaluate 获取父元素类名
                try:
                    parent_class = next_btn.evaluate("el => el.parentElement.className")
                    if parent_class and "disabled" in parent_class:
                        print("下一页按钮已禁用 (li.disabled)，扫描结束。")
                        break
                except:
                    pass

                try:
                    # 获取当前URL以便对比
                    current_url = page.url
                    with page.expect_navigation(timeout=10000):
                        next_btn.click()
                    
                    # 再次确认是否真的跳转了
                    if page.url == current_url:
                        print("点击下一页后URL未变化，可能已到末尾。")
                        break
                        
                    page_num += 1
                except Exception as e:
                    print(f"翻页操作异常: {e}")
                    # 尝试强制跳转到下一页URL (如果能推测出规律)
                    # 通常 URL 里有 page=X 参数
                    if "page=" in page.url:
                        new_url = re.sub(r'page=\d+', f'page={page_num+1}', page.url)
                        print(f"尝试强制跳转到: {new_url}")
                        try:
                            page.goto(new_url)
                            page_num += 1
                        except:
                             print("强制跳转失败，停止扫描。")
                             break
                    else:
                        break
            else:
                print("未找到下一页按钮，扫描结束。")
                break

        print(f"\n扫描结束，共发现 {len(ids_to_process)} 个新商品需要抓取。")
        
        # --- 第二阶段：批量抓取详情 ---
        if ids_to_process:
            print("\n=== 第二阶段：批量抓取详情 ===")
            processed_count = 0
            
            for goods_id in ids_to_process:
                processed_count += 1
                print(f"[{processed_count}/{len(ids_to_process)}] 正在处理 ID: {goods_id}")
                
                max_retries = 3
                for retry in range(max_retries):
                    try:
                        # 每次重试前确保上下文有效
                        if context.pages and len(context.pages) > 0 and context.pages[0].is_closed():
                             print("  上下文已关闭，尝试重建...")
                             # 这里比较复杂，简单起见只重试页面
                        
                        detail_page = context.new_page()
                        detail_url = f"https://szguokuai.zlj.xyzulin.top/web/index.php?c=site&a=entry&m=ewei_shopv2&do=web&r=goods.edit&id={goods_id}&goodsfrom=sale&page=1"
                        
                        # 设置超时
                        detail_page.set_default_timeout(15000)
                        
                        try:
                            detail_page.goto(detail_url, timeout=20000)
                        except Exception as nav_err:
                            print(f"  导航失败 ({retry+1}/{max_retries}): {nav_err}")
                            detail_page.close()
                            continue

                        try:
                            detail_page.wait_for_selector("#goodsname", state="visible", timeout=15000)
                            
                            goods_name = detail_page.input_value("#goodsname")
                            
                            short_title = ""
                            short_title_selector = "#tab_basic > div > div:nth-child(1) > div.region-goods-right.col-sm-10 > div:nth-child(3) > div > input"
                            if detail_page.query_selector(short_title_selector):
                                short_title = detail_page.input_value(short_title_selector)
                            
                            # 抓取分类信息
                            cate1 = ""
                            cate2 = ""
                            cate3 = ""
                            
                            def get_cate_text(p, sel):
                                try:
                                    txt = p.eval_on_selector(sel, "el => el.options[el.selectedIndex] ? el.options[el.selectedIndex].text : ''").strip()
                                    return "" if "请选择" in txt else txt
                                except:
                                    return ""

                            # 等待分类加载 (简单等待)
                            try:
                                detail_page.wait_for_selector("#cate1", state="attached", timeout=5000)
                            except: pass

                            cate1 = get_cate_text(detail_page, "#cate1")
                            cate2 = get_cate_text(detail_page, "#cate2")
                            cate3 = get_cate_text(detail_page, "#cate3")
                            
                            print(f"  分类: {cate1} | {cate2} | {cate3}")
                            
                            # --- SKU 表格抓取优化 ---
                            sku_table = detail_page.query_selector("#options > table")
                            
                            if sku_table:
                                # 1. 获取动态表头
                                headers = []
                                try:
                                    ths = sku_table.query_selector_all("thead th")
                                    for th in ths:
                                        h_text = th.inner_text().strip()
                                        if h_text:
                                            headers.append(h_text)
                                except:
                                    pass
                                
                                # 更新主表头顺序
                                update_master_headers(master_sku_headers, headers)
                                
                                # 2. 获取所有数据行
                                # 使用新的解析函数
                                sku_rows = parse_sku_table(sku_table, master_sku_headers)
                                
                                for row_data in sku_rows:
                                    # 补全基础信息
                                    row_data["ID"] = goods_id
                                    row_data["商品名称"] = goods_name
                                    row_data["短标题"] = short_title
                                    row_data["1级分类"] = cate1
                                    row_data["2级分类"] = cate2
                                    row_data["3级分类"] = cate3
                                    all_sku_rows.append(row_data)
                            else:
                                # 无SKU表格，仅保存基本信息
                                all_sku_rows.append({
                                    "ID": goods_id,
                                    "商品名称": goods_name,
                                    "短标题": short_title,
                                    "1级分类": cate1,
                                    "2级分类": cate2,
                                    "3级分类": cate3
                                })
                                
                            # 成功，跳出重试循环
                            detail_page.close()
                            break
                                
                        except Exception as e:
                            print(f"  抓取详情异常 ({retry+1}/{max_retries}): {e}")
                            detail_page.close()
                            if "Target page, context or browser has been closed" in str(e):
                                # 这种错误可能需要稍作等待
                                time.sleep(2)
                            continue
                        
                    except Exception as e:
                        print(f"  页面操作严重错误 ({retry+1}/{max_retries}): {e}")
                        if 'detail_page' in locals():
                            try: detail_page.close() 
                            except: pass
                        time.sleep(1)
        
        browser.close()

    if not all_sku_rows:
        print("没有抓取到任何SKU数据。")
        return

    print("开始整理数据并保存...")
    df = pd.DataFrame(all_sku_rows)
    
    # 调整列顺序
    # 1. 基础列
    base_cols = ["ID", "商品名称", "短标题", "1级分类", "2级分类", "3级分类", "SKU"]
    
    # 2. 动态列 (数据列)
    # 我们希望排除掉已经是 specs 的列，只保留数据列
    # 但由于我们在 parse_sku_table 里已经把 specs 合并成了 "SKU"，
    # 所以剩下的列应该就是数据列 + 基础列
    # 不过为了保险，我们还是根据 master_sku_headers 来排序剩下的列
    
    # 找出所有在 master_sku_headers 中出现过，且在 df 中的列 (且不是 SKU)
    # 注意：master_sku_headers 里包含原本的规格列名，但现在它们已经被合并了
    # 所以我们需要区分哪些是数据列
    
    # 简单的做法：把剩下的列都作为数据列，按它们在 master_sku_headers 出现的相对顺序排序
    remaining_cols = [c for c in df.columns if c not in base_cols]
    
    # 排序 remaining_cols
    def sort_key(col_name):
        # 1. 优先处理 "X天租金" 列，按天数数值排序
        rent_match = re.match(r"(\d+)天租金", col_name)
        if rent_match:
            days = int(rent_match.group(1))
            return (1, days)
        
        # 2. 处理特定已知列，固定顺序
        priority_cols = ["编号", "库存"]
        if col_name in priority_cols:
             return (0, priority_cols.index(col_name))
             
        post_cols = ["市场价", "押金", "购买价", "采购价"]
        if col_name in post_cols:
            return (2, post_cols.index(col_name))
        
        # 3. 其他列回退到 master_sku_headers 顺序
        if col_name in master_sku_headers:
            return (3, master_sku_headers.index(col_name))
            
        return (4, col_name)
        
    remaining_cols.sort(key=sort_key)
    
    final_cols = base_cols + remaining_cols
    
    # 确保所有列都在 df 中 (防止某些 base_cols 缺失)
    for col in final_cols:
        if col not in df.columns:
            df[col] = ""
            
    df = df.reindex(columns=final_cols)

    # 保存并设置样式
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 写入表头
        for c_idx, col_name in enumerate(final_cols, 1):
            ws.cell(row=1, column=c_idx, value=col_name)
            
        # 写入数据并设置背景色
        # 使用 PatternFill
        fill_colors = ["E0F7FA", "F3E5F5"] # 浅蓝, 浅紫
        current_fill_idx = 0
        last_goods_id = None
        
        # openpyxl 的 row 从 1 开始，header 是 row 1
        # dataframe rows
        rows = dataframe_to_rows(df, index=False, header=False)
        
        for r_idx, row in enumerate(rows, 2):
            # 获取当前行的 ID (假设 ID 在第一列)
            # row 是一个 list
            current_id = row[0] # ID 列
            
            if current_id != last_goods_id:
                current_fill_idx = (current_fill_idx + 1) % len(fill_colors)
                last_goods_id = current_id
                
            fill = PatternFill(start_color=fill_colors[current_fill_idx], end_color=fill_colors[current_fill_idx], fill_type="solid")
            
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.fill = fill

        wb.save(OUTPUT_FILE)
        print(f"数据已保存到 {OUTPUT_FILE}")
        
    except Exception as e:
        print(f"保存 Excel 失败: {e}")
        # 降级保存
        df.to_excel(OUTPUT_FILE, index=False)
        print("已使用普通模式保存。")

    # 保存处理过的 ID
    successful_ids = list(set([str(row["ID"]) for row in all_sku_rows]))
    save_new_ids(successful_ids)
    print("完成。")

def parse_sku_table(sku_table, master_headers):
    """
    解析 SKU 表格，处理 rowspan，合并规格列，提取数据列。
    返回: list of dict (rows)
    """
    rows = []
    try:
        # 1. 获取表头
        thead = sku_table.query_selector("thead")
        if not thead: return []
        
        headers = []
        ths = thead.query_selector_all("th")
        for th in ths:
            headers.append(th.inner_text().strip())
            
        # 更新主表头记录
        update_master_headers(master_headers, headers)
        
        # 2. 构建网格 (处理 rowspan)
        tbody = sku_table.query_selector("tbody")
        if not tbody: return []
        
        tr_elements = tbody.query_selector_all("tr")
        if not tr_elements: return []
        
        # 初始化网格
        # 这是一个动态增长的网格，row_idx -> col_idx -> {text, element}
        grid = {} 
        
        current_row_idx = 0
        for tr in tr_elements:
            tds = tr.query_selector_all("td")
            current_col_idx = 0
            
            # 确保当前行存在
            if current_row_idx not in grid:
                grid[current_row_idx] = {}
                
            for td in tds:
                # 跳过已被占用的位置
                while current_col_idx in grid[current_row_idx]:
                    current_col_idx += 1
                
                # 获取属性
                rowspan = 1
                colspan = 1
                try:
                    rs = td.get_attribute("rowspan")
                    if rs: rowspan = int(rs)
                except: pass
                
                try:
                    cs = td.get_attribute("colspan")
                    if cs: colspan = int(cs)
                except: pass
                
                # 提取数据
                # 优先找 input
                val = ""
                is_data_col = False
                
                input_el = td.query_selector("input:not([type='hidden'])")
                if input_el:
                    val = input_el.input_value()
                    is_data_col = True
                else:
                    select_el = td.query_selector("select")
                    if select_el:
                        val = select_el.input_value()
                        is_data_col = True # select 通常也是数据设定
                    else:
                        val = td.inner_text().strip()
                
                cell_info = {
                    "value": val,
                    "is_data": is_data_col
                }
                
                # 填充网格
                for r in range(rowspan):
                    for c in range(colspan):
                        target_row = current_row_idx + r
                        target_col = current_col_idx + c
                        
                        if target_row not in grid:
                            grid[target_row] = {}
                        
                        grid[target_row][target_col] = cell_info
                        
                current_col_idx += colspan
                
            current_row_idx += 1
            
        # 3. 从网格生成数据行
        num_cols = len(headers)
        
        # 识别哪些列是“数据列”，哪些是“规格列”
        # 启发式规则：
        # - 列名包含 "库存", "编号", "租金", "价格", "重量", "编码" -> 数据列
        # - 单元格包含 input -> 数据列 (在 cell_info 中已标记)
        # - 其他 -> 规格列
        
        spec_col_indices = []
        data_col_indices = []
        
        for idx, h in enumerate(headers):
            h_lower = h.lower()
            if any(k in h_lower for k in ["库存", "编号", "租金", "价格", "重量", "编码", "id"]):
                data_col_indices.append(idx)
            else:
                # 进一步检查该列在所有行中是否包含 input
                has_input = False
                for r in range(current_row_idx):
                    if r in grid and idx in grid[r] and grid[r][idx]["is_data"]:
                        has_input = True
                        break
                
                if has_input:
                    data_col_indices.append(idx)
                else:
                    spec_col_indices.append(idx)

        # 生成每一行的数据
        for r in range(current_row_idx):
            if r not in grid: continue
            
            row_data = {}
            specs = []
            
            # 收集规格
            for c in spec_col_indices:
                if c in grid[r]:
                    val = grid[r][c]["value"]
                    if val:
                        # 加上表头名称，格式：表头：值
                        header_name = headers[c] if c < len(headers) else ""
                        if header_name:
                            specs.append(f"{header_name}：{val}")
                        else:
                            specs.append(val)
            
            row_data["SKU"] = "|".join(specs)
            
            # 收集数据
            for c in data_col_indices:
                if c in grid[r]:
                    header_name = headers[c] if c < len(headers) else f"Col_{c}"
                    row_data[header_name] = grid[r][c]["value"]
            
            rows.append(row_data)
            
    except Exception as e:
        print(f"解析 SKU 表格出错: {e}")
        import traceback
        traceback.print_exc()
        
    return rows

if __name__ == "__main__":
    run_scraping()
